from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import openpyxl
import requests
from io import BytesIO
import zipfile
from datetime import datetime
import json
from dotenv import load_dotenv
import os

# Load environment variables
load_dotenv()

app = Flask(__name__)
CORS(app)

# Global variable for storing the original file
original_excel_file = None

def format_date(date_value):
    if not date_value:
        return ""
    try:
        if isinstance(date_value, datetime):
            return date_value.strftime("%d/%m/%Y")
        elif isinstance(date_value, str):
            date_value = date_value.replace('//', '/')
            day, month, year = date_value.split('/')
            year = int(year)
            if year < 2000:
                year += 2000
            return f"{int(day):02d}/{int(month):02d}/{year}"
    except Exception as e:
        print(f"Error formatting date: {e}")
        return str(date_value)
    return str(date_value)

def download_master_file():
    url = os.getenv('MASTER_FILE_URL')
    response = requests.get(url)
    return BytesIO(response.content)

@app.route('/')
def home():
    return 'Hello, World!'

@app.route('/generate-json', methods=['POST'])
def generate_json():
    global original_excel_file

    if not original_excel_file:
        return jsonify({'error': 'No original file found'}), 400

    try:
        original_excel_file.seek(0)
        wb = openpyxl.load_workbook(original_excel_file)

        formatted_data = {
            'customer_profile': {
                'company_name': '',
                'location': '',
                'official_account_name': '',
                'basic_id': '',
                'provider_id': '',
                'provider_name': '',
                'chanel_id': '',
                'chanel_secret': '',
                'contact_name': '',
                'contact_email': '',
                'campaign_start': '',
                'campaign_end': '',
                'special_request': '',
                'submit_date': ''
            },
            'devices': [],
            'info': []
        }

        # Read CustomerProfile sheet
        if 'CustomerProfile' in wb.sheetnames:
            sheet = wb['CustomerProfile']

            key_mapping = {
                "Company Name": "company_name",
                "Location": "location",
                "LINE Official Account name": "official_account_name",
                "LINE Basic ID / Premium ID": "basic_id",
                "LINE Provider ID": "provider_id",
                "LINE Provider Name": "provider_name",
                "LINE Chanel ID": "chanel_id",
                "LINE Chanel Secret": "chanel_secret",
                "Contact Name": "contact_name",
                "Contact Email": "contact_email",
                "Campaign Start": "campaign_start",
                "Campaign End": "campaign_end",
                "Special Request": "special_request",
                "Submit Date": "submit_date"
            }

            for row in range(1, sheet.max_row + 1):
                key = sheet[f'A{row}'].value
                value = sheet[f'B{row}'].value

                if not key or key == "ข้อมูลบริษัทและบัญชี LINE OA":
                    continue

                if key in key_mapping:
                    field = key_mapping[key]
                    if isinstance(value, str):
                        value = value.strip()
                    elif value is None:
                        value = ""

                    if key in ["Campaign Start", "Campaign End", "Submit Date"]:
                        value = format_date(value) if value else ""
                    elif key in ["LINE Provider ID", "LINE Chanel ID"]:
                        value = str(value) if value else ""

                    formatted_data['customer_profile'][field] = value

        # Read Devices sheet
        if 'Devices' in wb.sheetnames:
            sheet = wb['Devices']

            for row in range(2, sheet.max_row + 1):
                if not sheet[f'B{row}'].value:
                    break

                banner_type = str(sheet[f'J{row}'].value).strip() if sheet[f'J{row}'].value else ""
                if banner_type.startswith('='):
                    banner_type = "Link URL"

                device = {
                    'no': row - 1,
                    'hwid': str(sheet[f'B{row}'].value).strip() if sheet[f'B{row}'].value else "",
                    'banner_title_th': str(sheet[f'C{row}'].value).strip() if sheet[f'C{row}'].value else "",
                    'banner_message_th': str(sheet[f'E{row}'].value).strip() if sheet[f'E{row}'].value else "",
                    'banner_title_en': str(sheet[f'G{row}'].value).strip() if sheet[f'G{row}'].value else "",
                    'banner_message_en': str(sheet[f'I{row}'].value).strip() if sheet[f'I{row}'].value else "",
                    'banner_type': banner_type,
                    'banner_url': str(sheet[f'L{row}'].value).strip() if sheet[f'L{row}'].value else "",
                    'location': str(sheet[f'M{row}'].value).strip() if sheet[f'M{row}'].value else ""
                }
                formatted_data['devices'].append(device)

        # Read Info sheet
        if 'Info' in wb.sheetnames:
            sheet = wb['Info']
            for row in range(2, sheet.max_row + 1):
                if sheet[f'A{row}'].value:
                    value = str(sheet[f'A{row}'].value).strip()
                    formatted_data['info'].append({
                        'type': value
                    })
                else:
                    break

        return jsonify(formatted_data), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/update-directus', methods=['POST'])
def update_directus():
    try:
        if not original_excel_file:
            return jsonify({'error': 'No original file found'}), 400

        # Get JSON data from generate_json function
        response = generate_json()
        if response[1] != 200:  # Check if generate_json was successful
            return response

        data = response[0].json  # Get the JSON data

        # Send POST request to AWS API Gateway
        aws_api_url = os.getenv('AWS_API_URL')
        headers = {
            'Content-Type': 'application/json',
            'Accept': 'application/json'
        }

        print("Sending data:", json.dumps(data, indent=2, ensure_ascii=False))  # Debug print

        response = requests.post(
            aws_api_url,
            json=data,
            headers=headers
        )

        print("Response:", response.status_code, response.text)  # Debug print

        if response.status_code == 200:
            return jsonify(response.json()), 200
        else:
            return jsonify({
                'error': 'Failed to update Directus',
                'status': response.status_code,
                'response': response.text
            }), response.status_code

    except Exception as e:
        print(f"Error in update_directus: {str(e)}")  # Debug print
        return jsonify({'error': str(e)}), 500

@app.route('/upload', methods=['POST'])
def upload_file():
    global original_excel_file

    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400

    if file and file.filename.endswith('.xlsx'):
        try:
            file_content = file.read()
            original_excel_file = BytesIO(file_content)
            working_file = BytesIO(file_content)
            wb = openpyxl.load_workbook(working_file)

            # Process the file and create ZIP
            memory_zip = BytesIO()
            with zipfile.ZipFile(memory_zip, 'w') as zf:
                master_template = download_master_file()

                # Read CustomerProfile data
                customer_sheet = wb['CustomerProfile']
                customer_data = {}
                for row in range(1, customer_sheet.max_row + 1):
                    key = customer_sheet[f'A{row}'].value
                    value = customer_sheet[f'B{row}'].value
                    if key and isinstance(key, str):
                        if isinstance(value, str):
                            value = value.strip()
                        customer_data[key.strip()] = value

                # Read Devices data
                devices_sheet = wb['Devices']
                devices_data = []
                for row in range(2, devices_sheet.max_row + 1):
                    if not devices_sheet[f'B{row}'].value:
                        break
                    devices_data.append({
                        'id': devices_sheet[f'B{row}'].value,
                        'url': devices_sheet[f'L{row}'].value,
                        'column_c': devices_sheet[f'C{row}'].value,
                        'column_e': devices_sheet[f'E{row}'].value,
                        'column_g': devices_sheet[f'G{row}'].value,
                        'column_i': devices_sheet[f'I{row}'].value,
                        'location': devices_sheet[f'M{row}'].value or '[Official name of the Location owner]'
                    })

                # Create files for each device
                for index, device in enumerate(devices_data, 1):
                    output = create_output_file(master_template, customer_data, device)
                    shell_e19 = customer_data.get('LINE Official Account name', '')
                    device_id = device['id']
                    filename = f"[{shell_e19}]{index}[{device_id}]_TH-LINE Beacon Banner_Stay event Application Form-v1-1.xlsx"
                    zf.writestr(filename, output.getvalue())

            memory_zip.seek(0)
            return send_file(
                memory_zip,
                mimetype='application/zip',
                as_attachment=True,
                download_name='updated_files.zip'
            )

        except Exception as e:
            return jsonify({'error': str(e)}), 500

    return jsonify({'error': 'Invalid file format'}), 400

def create_output_file(master_template, customer_data, device_data):
    wb = openpyxl.load_workbook(master_template)
    sheet = wb.active

    # Map customer data to cells
    mapping = {
        'Location': 'D7',
        'Company Name': 'D9',
        'LINE Official Account name': 'E19',
        'LINE Basic ID / Premium ID': 'K19',
        'LINE Provider ID': 'D17',
        'LINE Provider Name': 'D18',
        'Special Request': 'D48'
    }

    for key, cell in mapping.items():
        if key in customer_data:
            value = customer_data[key]
            if isinstance(value, str):
                value = value.strip()
            sheet[cell] = value

    # Handle Submit Date separately
    if 'Submit Date' in customer_data:
        try:
            formatted_date = format_date(customer_data['Submit Date'])
            sheet['K60'] = formatted_date
        except Exception as e:
            print(f"Error formatting Submit Date: {e}")

    # Update Device data
    sheet['D23'] = device_data['id']
    sheet['D39'] = device_data['id']
    sheet['D36'] = device_data['url']
    sheet['F41'] = device_data['column_c']
    sheet['F42'] = device_data['column_e']
    sheet['F44'] = device_data['column_g']
    sheet['F45'] = device_data['column_i']

    # Add location owner information to D24
    location = device_data.get('location', '').strip()
    sheet['D24'] = location if location else '[Official name of the Location owner]'

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

if __name__ == '__main__':
    app.run(debug=True)